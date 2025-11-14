from datetime import date, datetime
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook

# ==== Paramètres de l'obligation ====
FACE = 100.0
COUPON = 0.0475           # 4.75% / an
FREQ = 1                  # annuel
ISSUE_DATE = date(2023, 3, 14)
MATURITY_DATE = date(2030, 3, 14)
COUPON_DAY = 8            # 08/10/X
COUPON_MONTH = 10
DAYCOUNT = "ACT/ACT"      # ACT/ACT (ISDA simplifié)
IS_PRICE_CLEAN = True     # True: B2 est clean -> on ajoute l'accrued pour obtenir le dirty

# ========= utilitaires =========
def is_leap(y:int)->bool:
    return (y % 4 == 0 and y % 100 != 0) or (y % 400 == 0)

def yearfrac_act_act(d1:date, d2:date) -> float:
    if d2 <= d1:
        return 0.0
    y1, y2 = d1.year, d2.year
    if y1 == y2:
        denom = 366 if is_leap(y1) else 365
        return (d2 - d1).days / denom
    # d1 -> 31/12/y1
    end_y1 = date(y1, 12, 31)
    denom1 = 366 if is_leap(y1) else 365
    part1 = (end_y1 - d1).days + 1
    # années pleines au milieu
    full_years = sum(1.0 for y in range(y1 + 1, y2))
    # 1/1/y2 -> d2
    start_y2 = date(y2, 1, 1)
    denom2 = 366 if is_leap(y2) else 365
    part2 = (d2 - start_y2).days + 1
    return part1/denom1 + full_years + part2/denom2

def yearfrac(d1:date, d2:date) -> float:
    return yearfrac_act_act(d1, d2)

def anchor_on_oct8(y:int) -> date:
    return date(y, COUPON_MONTH, COUPON_DAY)

def build_schedule(issue:date, maturity:date):
    """ Paiements réguliers au 08/10, puis stub final à la maturité si besoin. """
    pays = []
    y = issue.year
    first_anchor = anchor_on_oct8(y)
    if first_anchor <= issue:
        first_anchor = anchor_on_oct8(y + 1)
    d = first_anchor
    while d <= maturity:
        pays.append(d)
        d = date(d.year + 1, COUPON_MONTH, COUPON_DAY)
    if not pays or pays[-1] != maturity:
        pays.append(maturity)  # stub final si maturité != 08/10
    return pays

def accrued_interest(asof:date) -> float:
    """ Intérêt couru au 'asof' (pro rata ACT/ACT entre dernier 08/10 <= asof et le prochain 08/10) """
    if asof <= ISSUE_DATE or asof >= MATURITY_DATE:
        return 0.0
    prev = anchor_on_oct8(asof.year)
    if prev > asof:
        prev = anchor_on_oct8(asof.year - 1)
    nxt = anchor_on_oct8(prev.year + 1)
    period_len = yearfrac(prev, nxt)
    if period_len <= 0:
        return 0.0
    accrual = max(0.0, min(1.0, yearfrac(prev, asof) / period_len))
    coupon_full = COUPON * FACE / FREQ
    return coupon_full * accrual

def pv_from_y(asof:date, y:float) -> float:
    """ Valeur actuelle (dirty) des flux futurs au 'asof' pour un YTM nominal annualisé y (comp. annuel). """
    if asof >= MATURITY_DATE:
        return 0.0
    pays = build_schedule(ISSUE_DATE, MATURITY_DATE)
    c_full = COUPON * FACE / FREQ
    dirty = 0.0
    for d in pays:
        if d <= asof:
            continue
        # coupon normal ou coupon de stub final si d == maturité et != 08/10
        if d == MATURITY_DATE and (d.month != COUPON_MONTH or d.day != COUPON_DAY):
            prev_anchor = anchor_on_oct8(d.year - 1)
            next_anchor = anchor_on_oct8(d.year)
            frac = yearfrac(prev_anchor, d) / yearfrac(prev_anchor, next_anchor)
            coup = c_full * frac
        else:
            coup = c_full
        cf = coup + (FACE if d == MATURITY_DATE else 0.0)
        t = yearfrac(asof, d)
        dirty += cf / (1.0 + y) ** t
    return dirty

def ytm_from_price(asof:date, price:float, is_price_clean:bool=True,
                   tol:float=1e-10, max_iter:int=200) -> float:
    """ Résout le YTM nominal (comp. annuel) à partir du prix (clean ou dirty). """
    dirty_target = price + accrued_interest(asof) if is_price_clean else price
    # bornes robustes pour la bissection
    low, high = -0.99, 10.0  # -99% à 1000%
    f_low = pv_from_y(asof, low) - dirty_target
    f_high = pv_from_y(asof, high) - dirty_target
    if f_low * f_high > 0:
        raise RuntimeError("Impossible d'encadrer la racine: vérifie inputs (prix, dates).")
    for _ in range(max_iter):
        mid = 0.5 * (low + high)
        f_mid = pv_from_y(asof, mid) - dirty_target
        if abs(f_mid) < tol:
            return mid
        if f_low * f_mid < 0:
            high, f_high = mid, f_mid
        else:
            low, f_low = mid, f_mid
    return 0.5 * (low + high)

def ytm_effectif_annuel(ytm_nominal:float) -> float:
    """ Convertit un YTM nominal (comp. annuel) en taux effectif annuel. (identique ici car FREQ=1) """
    return (1.0 + ytm_nominal) ** 1 - 1.0

# ========= lecture/écriture Excel =========
def write_ytm_into_excel(path:str, sheet_name:str=None, row:int=2, col_A:int=1, col_B:int=2, col_C:int=3, col_D:int=4):
    """
    A{row} = date de valorisation (Excel ou 'DD/MM/YYYY')
    B{row} = prix (par défaut: clean)
    Écrit YTM nominal en C{row} et taux effectif en D{row}.
    """
    wb = load_workbook(path)
    ws = wb[sheet_name] if sheet_name else wb.active

    asof_raw = ws.cell(row=row, column=col_A).value
    px_raw = ws.cell(row=row, column=col_B).value

    # parse date
    if isinstance(asof_raw, datetime):
        asof = asof_raw.date()
    elif isinstance(asof_raw, date):
        asof = asof_raw
    else:
        try:
            asof = datetime.strptime(str(asof_raw), "%Y-%m-%d").date()
        except:
            asof = datetime.strptime(str(asof_raw), "%d/%m/%Y").date()

    price = float(px_raw)

    ytm = ytm_from_price(asof, price, is_price_clean=IS_PRICE_CLEAN)
    eff = ytm_effectif_annuel(ytm)

    ws.cell(row=row, column=col_C).value = ytm
    ws.cell(row=row, column=col_C).number_format = "0.0000%"
    ws.cell(row=row, column=col_D).value = eff
    ws.cell(row=row, column=col_D).number_format = "0.0000%"

    wb.save(path)
    print(f"Écrit: YTM={ytm:.8f} (C{row}), Effectif={eff:.8f} (D{row})")

write_ytm_into_excel('/Users/beldjenna/Desktop/XS2597110027=TX Overview.csv',sheet_name=None,rows=2)

# ======== Exemple d'utilisation ========
# write_ytm_into_excel("ton_fichier.xlsx", sheet_name=None, row=2)
